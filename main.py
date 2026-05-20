# ==========================================
# main.py — pipeline orchestrator
# Run: python main.py --disease diabetes
# ==========================================

import os
import argparse
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR, MAX_DRUGS
from disease_validator import validate_disease_input
from fetcher import fetch_all
from ml import (
    compute_topological_indices,
    run_correlation,
    run_ml_qspr,
    compute_shap,
    apply_drug_filters,
)

os.makedirs(OUTPUT_DIR, exist_ok=True)


# ==========================================
# EXCEL EXPORT — all results in one workbook
# ==========================================
def _style_header(cell, bg="1F4E79"):
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _style_subheader(cell):
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color="2E75B6")
    cell.alignment = Alignment(horizontal="center", vertical="center")

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)


def export_excel(disease, df, corr, ml_results, top_models, shap_summaries, out_path):
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"] = f"MolecuLens QSPR Pipeline — {disease.title()}"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E79", name="Arial")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Disease"; ws["B3"] = disease.title()
    ws["A4"] = "Drugs analysed"; ws["B4"] = len(df)
    ws["A5"] = "Properties per drug"; ws["B5"] = len(df.columns)

    lipinski = int(df["Lipinski_Pass"].sum()) if "Lipinski_Pass" in df.columns else "N/A"
    veber    = int(df["Veber_Pass"].sum())    if "Veber_Pass"    in df.columns else "N/A"
    pains    = int(df["PAINS_Pass"].eq(True).sum()) if "PAINS_Pass" in df.columns else "N/A"
    ws["A6"] = "Lipinski pass"; ws["B6"] = lipinski
    ws["A7"] = "Veber pass";    ws["B7"] = veber
    ws["A8"] = "PAINS clean";   ws["B8"] = pains

    if not top_models.empty:
        best = top_models.sort_values("R2_mean", ascending=False).iloc[0]
        best_r2 = round(float(best["R2_mean"]), 4)
        best_model_desc = f"{best['Model']} predicting {best['Property']}"
    else:
        best_r2 = "N/A"
        best_model_desc = "N/A (Insufficient data for ML)"

    ws["A9"]  = "Best R² overall"
    ws["B9"]  = best_r2
    ws["A10"] = "Best model"
    ws["B10"] = best_model_desc

    for row in ws["A3:A10"]:
        for cell in row:
            cell.font = Font(bold=True, name="Arial", size=10)
    for row in ws["B3:B10"]:
        for cell in row:
            cell.font = Font(name="Arial", size=10)

    ws["A12"] = "Sheets in this workbook:"
    ws["A12"].font = Font(bold=True, name="Arial", size=10)
    sheets_info = [
        ("Drug Data",        "All drugs with physicochemical, ADMET, topological and filter properties"),
        ("ML Results",       "R², MAE for all 4 models × 7 properties with k-fold CV"),
        ("Best Models",      "Best model per property with performance metrics"),
        ("SHAP Importance",  "Feature importance (SHAP values) per property"),
        ("Correlation",      "Pearson correlation matrix — indices vs properties"),
        ("Topological Index","Definition and formula for all 50 topological indices"),
    ]
    for i, (name, desc) in enumerate(sheets_info, start=13):
        ws[f"A{i}"] = name
        ws[f"B{i}"] = desc
        ws[f"A{i}"].font = Font(bold=True, color="2E75B6", name="Arial", size=10)
        ws[f"B{i}"].font = Font(name="Arial", size=10)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 55

    # ── Sheet 2: Drug Data ────────────────────────────────
    ws2 = wb.create_sheet("Drug Data")
    ws2.sheet_view.showGridLines = False
    headers = list(df.columns)
    for ci, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=ci, value=h)
        _style_header(cell)
    for ri, row in enumerate(df.itertuples(index=False), start=2):
        for ci, val in enumerate(row, start=1):
            c = ws2.cell(row=ri, column=ci, value=val if not isinstance(val, float) or val == val else None)
            c.font = Font(name="Arial", size=9)
            c.border = _thin_border()
            if ri % 2 == 0:
                c.fill = PatternFill("solid", start_color="EBF3FB")
    ws2.freeze_panes = "A2"
    _autofit(ws2)

    # ── Sheet 3: ML Results ───────────────────────────────
    ws3 = wb.create_sheet("ML Results")
    ws3.sheet_view.showGridLines = False
    ml_headers = ["Property", "Model", "R² Mean", "R² Std", "MAE Mean"]
    for ci, h in enumerate(ml_headers, start=1):
        _style_header(ws3.cell(row=1, column=ci, value=h))
    for ri, row in enumerate(ml_results.itertuples(index=False), start=2):
        vals = [row.Property, row.Model,
                round(float(row.R2_mean), 4),
                round(float(row.R2_std), 4),
                round(float(row.MAE_mean), 4)]
        for ci, val in enumerate(vals, start=1):
            c = ws3.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", size=10)
            c.border = _thin_border()
            if ri % 2 == 0:
                c.fill = PatternFill("solid", start_color="EBF3FB")
    ws3.freeze_panes = "A2"
    _autofit(ws3)

    # ── Sheet 4: Best Models ──────────────────────────────
    ws4 = wb.create_sheet("Best Models")
    ws4.sheet_view.showGridLines = False
    for ci, h in enumerate(ml_headers, start=1):
        _style_header(ws4.cell(row=1, column=ci, value=h))
    for ri, row in enumerate(top_models.itertuples(index=False), start=2):
        vals = [row.Property, row.Model,
                round(float(row.R2_mean), 4),
                round(float(row.R2_std), 4),
                round(float(row.MAE_mean), 4)]
        for ci, val in enumerate(vals, start=1):
            c = ws4.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", size=10)
            c.border = _thin_border()
            if ri % 2 == 0:
                c.fill = PatternFill("solid", start_color="EBF3FB")
    ws4.freeze_panes = "A2"
    _autofit(ws4)

    # ── Sheet 5: SHAP Importance ──────────────────────────
    ws5 = wb.create_sheet("SHAP Importance")
    ws5.sheet_view.showGridLines = False
    ws5.cell(row=1, column=1, value="Property")
    ws5.cell(row=1, column=2, value="Topological Index")
    ws5.cell(row=1, column=3, value="Mean |SHAP|")
    ws5.cell(row=1, column=4, value="Rank")
    for c in [ws5.cell(row=1, column=i) for i in range(1,5)]:
        _style_header(c)
    ri = 2
    for prop, feat_vals in shap_summaries.items():
        sorted_feats = sorted(feat_vals.items(), key=lambda x: x[1], reverse=True)
        for rank, (feat, val) in enumerate(sorted_feats, start=1):
            ws5.cell(row=ri, column=1, value=prop).font   = Font(name="Arial", size=10)
            ws5.cell(row=ri, column=2, value=feat).font   = Font(name="Arial", size=10)
            ws5.cell(row=ri, column=3, value=round(val,6)).font = Font(name="Arial", size=10)
            ws5.cell(row=ri, column=4, value=rank).font  = Font(name="Arial", size=10)
            for ci in range(1,5):
                ws5.cell(row=ri, column=ci).border = _thin_border()
                if ri % 2 == 0:
                    ws5.cell(row=ri, column=ci).fill = PatternFill("solid", start_color="EBF3FB")
            ri += 1
    ws5.freeze_panes = "A2"
    _autofit(ws5)

    # ── Sheet 6: Correlation Matrix (Pearson + p-values + Spearman) ─
    ws6 = wb.create_sheet("Correlation")
    ws6.sheet_view.showGridLines = False

    # corr is now a dict with keys: pearson, pearson_p, spearman, spearman_p, significance, vif
    pearson_dict = corr.get("pearson", {}) if isinstance(corr, dict) else {}
    pearson_p_dict = corr.get("pearson_p", {}) if isinstance(corr, dict) else {}
    spearman_dict = corr.get("spearman", {}) if isinstance(corr, dict) else {}
    sig_dict = corr.get("significance", {}) if isinstance(corr, dict) else {}
    vif_dict = corr.get("vif", {}) if isinstance(corr, dict) else {}

    props   = list(list(pearson_dict.values())[0].keys()) if pearson_dict else []
    indices_list = list(pearson_dict.keys()) if pearson_dict else []

    # Section headers
    ws6.cell(row=1, column=1, value="PEARSON CORRELATION")
    ws6["A1"].font = Font(bold=True, color="1F4E79", name="Arial", size=11)
    if props:
        ws6.merge_cells(f"A1:{get_column_letter(len(props)+1)}1")

    ws6.cell(row=2, column=1, value="Index / Property")
    _style_header(ws6.cell(row=2, column=1))
    for ci, p in enumerate(props, start=2):
        _style_header(ws6.cell(row=2, column=ci, value=p))

    for ri, idx in enumerate(indices_list, start=3):
        _style_subheader(ws6.cell(row=ri, column=1, value=idx))
        for ci, p in enumerate(props, start=2):
            val = float(pearson_dict.get(idx, {}).get(p, 0))
            pval = float(pearson_p_dict.get(idx, {}).get(p, 1))
            sig  = sig_dict.get(idx, {}).get(p, "")
            cell_val = f"{val:.3f} {sig}"
            c = ws6.cell(row=ri, column=ci, value=cell_val)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="center")
            c.border = _thin_border()
            abs_val = abs(val)
            if val > 0:
                intensity = int(abs_val * 180)
                r2, g2, b2 = 255-intensity, 255, 255-intensity
            else:
                intensity = int(abs_val * 180)
                r2, g2, b2 = 255, 255-intensity, 255-intensity
            ws6.cell(row=ri, column=ci).fill = PatternFill("solid", start_color=f"{r2:02X}{g2:02X}{b2:02X}")

    # Spearman section
    spear_row = len(indices_list) + 5
    ws6.cell(row=spear_row, column=1, value="SPEARMAN CORRELATION")
    ws6[f"A{spear_row}"].font = Font(bold=True, color="1F4E79", name="Arial", size=11)
    if props:
        ws6.merge_cells(f"A{spear_row}:{get_column_letter(len(props)+1)}{spear_row}")
    ws6.cell(row=spear_row+1, column=1, value="Index / Property")
    _style_header(ws6.cell(row=spear_row+1, column=1))
    for ci, p in enumerate(props, start=2):
        _style_header(ws6.cell(row=spear_row+1, column=ci, value=p))
    for ri2, idx in enumerate(indices_list, start=spear_row+2):
        _style_subheader(ws6.cell(row=ri2, column=1, value=idx))
        for ci, p in enumerate(props, start=2):
            val = float(spearman_dict.get(idx, {}).get(p, 0))
            c = ws6.cell(row=ri2, column=ci, value=round(val,3))
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="center")
            c.border = _thin_border()

    # VIF section
    vif_row = spear_row + len(indices_list) + 4
    ws6.cell(row=vif_row, column=1, value="VARIANCE INFLATION FACTOR (VIF) — Multicollinearity")
    ws6[f"A{vif_row}"].font = Font(bold=True, color="1F4E79", name="Arial", size=11)
    ws6.merge_cells(f"A{vif_row}:C{vif_row}")
    ws6.cell(row=vif_row+1, column=1, value="Index")
    ws6.cell(row=vif_row+1, column=2, value="VIF")
    ws6.cell(row=vif_row+1, column=3, value="Interpretation")
    for ci in [1,2,3]: _style_header(ws6.cell(row=vif_row+1, column=ci))
    for ri3, (idx, vif_val) in enumerate(vif_dict.items(), start=vif_row+2):
        ws6.cell(row=ri3, column=1, value=idx).font = Font(name="Arial", size=10)
        ws6.cell(row=ri3, column=2, value=vif_val).font = Font(name="Arial", size=10)
        if vif_val is None: interp = "N/A"
        elif vif_val < 5:   interp = "Low multicollinearity"
        elif vif_val < 10:  interp = "Moderate multicollinearity"
        else:               interp = "High multicollinearity — consider removing"
        ws6.cell(row=ri3, column=3, value=interp).font = Font(name="Arial", size=10)
        for ci in [1,2,3]:
            ws6.cell(row=ri3, column=ci).border = _thin_border()
            if ri3 % 2 == 0:
                ws6.cell(row=ri3, column=ci).fill = PatternFill("solid", start_color="EBF3FB")

    # Legend
    leg_row = vif_row + len(vif_dict) + 3
    ws6.cell(row=leg_row, column=1, value="Significance: *** p<0.001   ** p<0.01   * p<0.05   ns = not significant")
    ws6[f"A{leg_row}"].font = Font(italic=True, name="Arial", size=9, color="666666")

    _autofit(ws6)

    # ── Sheet 7: Index Definitions ────────────────────────
    ws7 = wb.create_sheet("Topological Index")
    ws7.sheet_view.showGridLines = False
    def_headers = ["Index", "Full Name", "Type", "Formula / Description"]
    for ci, h in enumerate(def_headers, start=1):
        _style_header(ws7.cell(row=1, column=ci, value=h))
    definitions = [
        # Original degree-based
        ("M1",   "First Zagreb Index",           "Degree",         "Σ d(v)² over all vertices"),
        ("M2",   "Second Zagreb Index",          "Degree",         "Σ d(u)·d(v) over all edges"),
        ("ABC",  "Atom-Bond Connectivity",       "Degree",         "Σ √((d(u)+d(v)-2)/(d(u)·d(v))) over edges"),
        ("R",    "Randic Connectivity",          "Degree",         "Σ 1/√(d(u)·d(v)) over edges"),
        ("H",    "Harmonic Index",               "Degree",         "Σ 2/(d(u)+d(v)) over edges"),
        ("F",    "Forgotten Index",              "Degree",         "Σ d(v)³ over all vertices"),
        ("AZI",  "Augmented Zagreb Index",       "Degree",         "Σ (d(u)·d(v)/(d(u)+d(v)-2))³ over edges"),
        ("GA",   "Geometric-Arithmetic Index",   "Degree",         "Σ 2√(d(u)·d(v))/(d(u)+d(v)) over edges"),
        ("SC",   "Sum-Connectivity Index",       "Degree",         "Σ 1/√(d(u)+d(v)) over edges"),
        # New degree-based from image
        ("BM",   "Bi-Zagreb Index",              "Degree",         "Σ (d(u)+d(v)+d(u)·d(v)) over edges"),
        ("TM",   "Tri-Zagreb Index",             "Degree",         "Σ (d(u)²+d(v)²+d(u)·d(v)) over edges"),
        ("GH",   "Geometric-Harmonic Index",     "Degree",         "Σ √(d(u)·d(v))·(d(u)+d(v))/2 over edges"),
        ("GBM",  "Geometric Bi-Zagreb Index",    "Degree",         "Σ √(d(u)·d(v))/(d(u)+d(v)+d(u)·d(v)) over edges"),
        ("GTM",  "Geometric Tri-Zagreb Index",   "Degree",         "Σ √(d(u)·d(v))/(d(u)²+d(v)²+d(u)·d(v)) over edges"),
        ("HG",   "Harmonic-Geometric Index",     "Degree",         "Σ 2/(√(d(u)·d(v))·(d(u)+d(v))) over edges"),
        ("BMG",  "Bi Zagreb-Geometric Index",    "Degree",         "Σ (d(u)+d(v)+d(u)·d(v))/√(d(u)·d(v)) over edges"),
        ("BMH",  "Bi Zagreb-Harmonic Index",     "Degree",         "Σ (d(u)+d(v)+d(u)·d(v))·(d(u)+d(v))/2 over edges"),
        ("TMG",  "Tri Zagreb-Geometric Index",   "Degree",         "Σ (d(u)²+d(v)²+d(u)·d(v))/√(d(u)·d(v)) over edges"),
        ("TMH",  "Tri Zagreb-Harmonic Index",    "Degree",         "Σ (d(u)²+d(v)²+d(u)·d(v))·(d(u)+d(v))/2 over edges"),
        ("SDD",  "Symmetric Degree Division",    "Degree",         "Σ (d(u)²+d(v)²)/(d(u)·d(v)) over edges"),
        # Reverse-degree variants
        ("RM1",  "Reverse First Zagreb",         "Reverse-degree", "Σ rd(v)² where rd(v)=n+1-d(v)"),
        ("RM2",  "Reverse Second Zagreb",        "Reverse-degree", "Σ rd(u)·rd(v) over edges"),
        ("RABC", "Reverse ABC Index",            "Reverse-degree", "Σ √((rd(u)+rd(v)-2)/(rd(u)·rd(v))) over edges"),
        ("RR",   "Reverse Randic Index",         "Reverse-degree", "Σ 1/√(rd(u)·rd(v)) over edges"),
        ("RH",   "Reverse Harmonic Index",       "Reverse-degree", "Σ 2/(rd(u)+rd(v)) over edges"),
        ("RF",   "Reverse Forgotten Index",      "Reverse-degree", "Σ rd(v)³ over all vertices"),
        ("RGA",  "Reverse Geometric-Arithmetic", "Reverse-degree", "Σ 2√(rd(u)·rd(v))/(rd(u)+rd(v)) over edges"),
        ("RBM",  "Reverse Bi-Zagreb",            "Reverse-degree", "Σ (rd(u)+rd(v)+rd(u)·rd(v)) over edges"),
        ("RSDD", "Reverse Sym. Degree Division", "Reverse-degree", "Σ (rd(u)²+rd(v)²)/(rd(u)·rd(v)) over edges"),
        # Degree-sum variants
        ("DS1",  "Degree-Sum Zagreb 1",          "Degree-sum",     "Σ (d(u)+d(v))² over edges"),
        ("DS2",  "Degree-Sum Zagreb 2",          "Degree-sum",     "Σ (d(u)+d(v))·(d(u)+d(v)) over edges"),
        ("DSR",  "Degree-Sum Randic",            "Degree-sum",     "Σ 1/√(d(u)+d(v)) over edges"),
        ("DSH",  "Degree-Sum Harmonic",          "Degree-sum",     "Σ 2/(d(u)+d(v)) over edges"),
        ("DSGA", "Degree-Sum GA",                "Degree-sum",     "Σ 2√(d(u)·d(v))/(d(u)+d(v)) over edges"),
        # Distance-based
        ("W",    "Wiener Index",                 "Distance",       "Σ d(u,v) all pairs shortest-path distances"),
        ("J",    "Balaban J Index",              "Distance",       "(m/(m-n+2))·Σ 1/√(s(u)·s(v)) per edge"),
        ("Z",    "Hosoya Z Index",               "Distance",       "Total number of matchings in the molecular graph"),
        ("Sz",   "Szeged Index",                 "Distance",       "Σ n_u(e)·n_v(e) over edges"),
        ("GE",   "Graph Entropy",                "Information",    "-Σ p(d)·log₂(p(d)) Shannon entropy of degrees"),
        # NEW: Advanced distance-based Wiener variants (6)
        ("W_v",  "Vertex Wiener",                "Distance",       "Vertex Wiener index, equivalent to W"),
        ("W_e",  "Edge Wiener",                  "Distance",       "Edge Wiener index based on line graph distances"),
        ("W_ve", "Vertex-Edge Wiener",           "Distance",       "Mixed Vertex-Edge Wiener index"),
        ("Sz_v", "Vertex Szeged",                "Distance",       "Vertex Szeged index, equivalent to Sz"),
        ("Sz_e", "Edge Szeged",                  "Distance",       "Edge Szeged index based on edge partitioning"),
        ("Sz_ve","Vertex-Edge Szeged",           "Distance",       "Vertex-Edge Szeged index"),
        # NEW: Mostar indices (2)
        ("Mo_v", "Vertex Mostar",                "Distance",       "Vertex Mostar index measuring peripheral distance asymmetry"),
        ("Mo_e", "Edge Mostar",                  "Distance",       "Edge Mostar index measuring bond distance asymmetry"),
        # NEW: Special indices (3)
        ("PI",   "Padmakar-Ivan Index",          "Distance",       "PI index based on edge partition cuts"),
        ("Schultz","Schultz Index",              "Distance",       "Schultz molecular topological index: Σ (d_i + d_j)·d_ij"),
        ("Gutman", "Gutman Index",               "Distance",       "Gutman index with degree weightings: Σ (d_i · d_j)·d_ij"),
    ]
    for ri, row in enumerate(definitions, start=2):
        for ci, val in enumerate(row, start=1):
            c = ws7.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", size=10)
            c.border = _thin_border()
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if ri % 2 == 0:
                c.fill = PatternFill("solid", start_color="EBF3FB")
    ws7.column_dimensions["A"].width = 8
    ws7.column_dimensions["B"].width = 28
    ws7.column_dimensions["C"].width = 16
    ws7.column_dimensions["D"].width = 55
    ws7.freeze_panes = "A2"

    wb.save(out_path)
    print(f"    [✓] Excel workbook saved → {out_path}")


def run_pipeline(disease: str, max_drugs: int = MAX_DRUGS) -> dict:
    print(f"\n{'='*50}")
    print(f"  Drug Pipeline: {disease.upper()}")
    print(f"{'='*50}\n")

    df = fetch_all(disease, max_drugs)
    if df.empty:
        return {"error": f"No data found for '{disease}'."}

    df = apply_drug_filters(df)
    df = compute_topological_indices(df)
    try:
        corr_data = run_correlation(df)
    except Exception as e:
        print(f"    [!] Correlation failed: {e}")
        corr_data = {"pearson": {}, "pearson_p": {}, "spearman": {}, "spearman_p": {}, "significance": {}, "vif": {}}
    ml_results, best_models = run_ml_qspr(df)
    try:
        shap_summaries = compute_shap(df, best_models)
    except Exception as e:
        print(f"    [!] SHAP failed: {e} — continuing without SHAP")
        shap_summaries = {}

    if ml_results.empty:
        top_models = pd.DataFrame(columns=["Property","Model","R2_mean","R2_std","MAE_mean"])
    else:
        top_models = (
            ml_results.sort_values("R2_mean", ascending=False)
            .drop_duplicates("Property")
            .reset_index(drop=True)
        )

    csv_path      = os.path.join(OUTPUT_DIR, f"{disease}_results.csv")
    xlsx_path     = os.path.join(OUTPUT_DIR, f"{disease}_results.xlsx")
    results_cache = os.path.join(OUTPUT_DIR, f"{disease}_results_cache.json")

    df.to_csv(csv_path, index=False)

    export_excel(
        disease=disease,
        df=df,
        corr=corr_data,
        ml_results=ml_results,
        top_models=top_models,
        shap_summaries=shap_summaries,
        out_path=xlsx_path,
    )

    print(f"\n[*] Pipeline complete")
    print(f"    Drugs processed : {len(df)}")
    print(f"    Total columns   : {len(df.columns)}")
    print(f"    CSV  → {csv_path}")
    print(f"    XLSX → {xlsx_path}")

    result = {
        "disease":        disease,
        "drug_count":     len(df),
        "columns":        list(df.columns),
        "correlation":    corr_data,
        "ml_results":     ml_results.to_dict(orient="records"),
        "top_models":     top_models.to_dict(orient="records"),
        "shap":           shap_summaries,
        "lipinski_pass":  int(df["Lipinski_Pass"].fillna(False).astype(bool).sum()) if "Lipinski_Pass" in df.columns else 0,
        "veber_pass":     int(df["Veber_Pass"].fillna(False).astype(bool).sum()) if "Veber_Pass" in df.columns else 0,
        "pains_pass":     int(pd.Series(df["PAINS_Pass"].values).eq(True).sum()) if "PAINS_Pass" in df.columns else 0,
        "csv_path":       csv_path,
        "xlsx_path":      xlsx_path,
        "drugs_preview":  df.head(10).to_dict(orient="records"),
    }

    # Cache full results so repeat searches return instantly
    try:
        import json, math

        def _clean(obj):
            if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
                return None
            if isinstance(obj, dict): return {k: _clean(v) for k,v in obj.items()}
            if isinstance(obj, list): return [_clean(v) for v in obj]
            return obj

        result = _clean(result)
        with open(results_cache, "w") as f:
            json.dump(result, f, default=str)
        print(f"    Results cached → {results_cache}")
    except Exception as e:
        print(f"    [!] Could not cache results: {e}")

    return result


def main():
    parser = argparse.ArgumentParser(description="Drug QSPR Pipeline")
    parser.add_argument("--disease",   type=str, required=True)
    parser.add_argument("--max_drugs", type=int, default=MAX_DRUGS)
    parser.add_argument("--no_cache",  action="store_true")
    args = parser.parse_args()

    if args.no_cache:
        import config
        config.CACHE_ENABLED = False

    validation = validate_disease_input(args.disease)
    if not validation["valid"]:
        for e in validation["errors"]:
            print(f"[!] {e}")
        if validation["suggestions"]:
            print("Did you mean:", ", ".join(f'"{s}"' for s in validation["suggestions"]))
        return

    if validation["warnings"]:
        for w in validation["warnings"]:
            print(f"[!] Warning: {w}")
        confirm = input("Proceed anyway? (y/n): ").strip().lower()
        if confirm != "y":
            return
        validation = validate_disease_input(validation["disease"], force_proceed=True)

    run_pipeline(validation["disease"], args.max_drugs)


if __name__ == "__main__":
    main()
